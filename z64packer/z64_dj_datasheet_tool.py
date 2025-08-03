import json
import os
import openpyxl
import gdown

url = "https://docs.google.com/spreadsheets/d/1Yvgjex502cB_dVvvZm0a88aGL4WNFOm-5XvEbZLkWqI/export"
workbookPath = "z64packer/dj_datasheet.xlsx"
propertiesPath = 'z64packer/z64musicpacker.properties'
songsPath = 'z64packer/z64songs.json'

def processExcel():
  try:
    # Check if this is a Z64 repo
    if not os.path.exists(propertiesPath):
      print('This is not an Z64 repository | Missing z64musicpacker.properties file')
      return False
    
    # Get the repo name
    repoName = ""
    with open(propertiesPath, encoding='ISO-8859-1') as propertiesFile:
       properties = json.load(propertiesFile)
       repoName = properties["name"]

    # Download the datasheet
    gdown.download(url, workbookPath, fuzzy=True, format="xlsx")
    workbook = openpyxl.load_workbook(workbookPath)

    # Collect the full names of the converters
    print("------ GETTING CONVERTER INFORMATION ------")
    converters = {}
    informationSheet = workbook["Information"]
    for row in range(2, informationSheet.max_row):
      abbrName = informationSheet.cell(row=row, column=1).value
      fullName = informationSheet.cell(row=row, column=2).value

      # Skip any null values
      if abbrName is None or fullName is None: continue
      converters[abbrName] = fullName
      print("Found converter: " + abbrName + " - " + fullName)


    # Open up the database, since we are going to modify it
    with open(songsPath, 'r+', encoding='ISO-8859-1') as databaseFile:
      print("OPENING SONG DATABASE FILE")
      database = json.load(databaseFile)
      foundSongs = 0
      notFoundSongs = 0
      
      # Now process the worksheets!
      for sheetName in workbook.sheetnames:
        prefix = "------"
        if repoName == "Darunia's Joy": prefix = "DJ"
        if repoName == "Ganondorf's Organ": prefix = "GO"

        # Only process specific repos
        if not sheetName.startswith(prefix):
           print("Skipping sheet " + sheetName + "...")
           continue
        
        # Finally, open the sheet
        sheet = workbook[sheetName]
        print("------ PROCESSING SHEET: " + sheetName + " ------")

        # We allow an offset, because some sheets can have 2 tables of data
        for offset in [0, 6]:
          for row in range(2, sheet.max_row):
              game = sheet.cell(row=row, column=2+offset).value
              title = sheet.cell(row=row, column=3+offset).value
              progress = sheet.cell(row=row, column=4+offset).value
              author = sheet.cell(row=row, column=5+offset).value or ""
              sample = sheet.cell(row=row, column=6+offset).value or ""

              # If a track is WIP, we skip it
              if game is None or title is None or not isinstance(title, str) or progress != "Done": continue

              # If a track is a boolean, then it's probably because the name of the song is "True" or "False"
              if isinstance(title, bool): title = "True" if title else "False"

              # Now that we found our song, search for it in the database
              #print("Found song: " + game + " - " + title + " | " + author + " | " + sample)
              songIndex = None
              for i, song in enumerate(database):
                if compareTexts(song["game"], game) and compareTexts(song["song"], title):
                  songIndex = i
                  break
                
              # If we find it, update it!
              if songIndex is None:
                notFoundSongs += 1
                print("<<<< NOT FOUND IN DATABASE: " + game + " - " + title + " | " + author + " | " + sample)
                #print("normalized: " + normalize(game) + " - " + normalize(title))

              else:
                foundSongs += 1
                if author in converters: database[songIndex]["converters"] = [converters[author]]
                if sample.startswith("https://"): database[songIndex]["preview"] = sample
                #print("Found in database! Updating...")

      # Report findings
      print("FOUND SONGS: " + str(foundSongs))
      print("NOT FOUND SONGS: " + str(notFoundSongs))

      # Replace song database with this one
      databaseFile.seek(0)
      json.dump(database, databaseFile, indent=2)
      databaseFile.truncate()
    
    # If everything when ok, then return true
    return True
  
  # On exceptions, always print them
  except Exception as e:
    print(e)
    return False
  
  # Always try to delete the datasheet, so no evindence remains
  finally:
    if os.path.exists(workbookPath): os.remove(workbookPath)


def safe_list_index(iterable, value, default = None):
    for i, item in enumerate(iterable):
        if item == value:
            return i
    return default


# Methods for better finding names of songs...
def compareTexts(a, b):
   nA = normalize(a)
   nB = normalize(b)
   return nA in nB or nB in nA
 
def normalize(text):
   t = text.upper().replace("É", "E")
   r = [" ", "'", ":", "-", "/", ".", "(", ")", "!", ";"]
   for char in r:
      t = t.replace(char, "")
   return t


def main():
  result = processExcel()
  if result: print("Process completed succesfully!")
  else: print("An error occured")


if __name__ == '__main__':
  main()